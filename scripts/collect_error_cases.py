from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

# =========================
# 固定路径
# =========================
PROJECT_DIR = Path(r"D:\Users\30126\Desktop\project")
MODELS_DIR = PROJECT_DIR / "models"
OUTPUT_DIR = PROJECT_DIR / "error_case_analysis"
OUTPUT_DIR.mkdir(exist_ok=True)

# =========================
# 这里改成你的真实目录名
# 左边：论文里显示的数据集名
# 右边：models目录里的真实实验文件夹名
# =========================
BEST_MODEL_BY_DATASET = {
    "ChnSentiCorp": "bert_chnsenticorp",
    "OnlineShopping10Cats": "bert_bigru_online",
    "WeiboSenti100k": "roberta_weibo",
}

# =========================
# 模型目录名 -> 论文展示名
# 不够就继续补
# =========================
MODEL_NAME_MAP = {
    "bert_chnsenticorp": "BERT",
    "macbert_chnsenticorp": "MacBERT",
    "roberta_chnsenticorp": "RoBERTa-wwm-ext",
    "bert_bigru_chnsenticorp": "BERT+BiGRU",
    "bert_lora_chnsenticorp": "BERT+LoRA",

    "bert_online": "BERT",
    "macbert_online": "MacBERT",
    "roberta_online": "RoBERTa-wwm-ext",
    "bert_bigru_online": "BERT+BiGRU",
    "bert_lora_online": "BERT+LoRA",

    "bert_weibo": "BERT",
    "macbert_weibo": "MacBERT",
    "roberta_weibo": "RoBERTa-wwm-ext",
    "bert_bigru_weibo": "BERT+BiGRU",
    "bert_lora_weibo": "BERT+LoRA",
}


def detect_text_column(df: pd.DataFrame) -> str:
    candidates = ["text", "sentence", "content", "review", "comment"]
    for col in candidates:
        if col in df.columns:
            return col
    for col in df.columns:
        if df[col].dtype == "object":
            return col
    raise ValueError("找不到文本列，请检查 test_predictions.csv 的列名。")


def detect_true_label_column(df: pd.DataFrame) -> str:
    candidates = ["label", "true_label", "gold_label", "y_true"]
    for col in candidates:
        if col in df.columns:
            return col
    raise ValueError("找不到真实标签列。")


def detect_pred_label_column(df: pd.DataFrame) -> str:
    candidates = ["pred_label", "prediction", "pred", "y_pred", "pred_label_id"]
    for col in candidates:
        if col in df.columns:
            return col
    raise ValueError("找不到预测标签列。")


def detect_confidence_column(df: pd.DataFrame) -> str | None:
    candidates = ["confidence", "score", "prob", "probability"]
    for col in candidates:
        if col in df.columns:
            return col
    return None


def normalize_label(x):
    try:
        xi = int(float(x))
        if xi == 0:
            return "负向"
        if xi == 1:
            return "正向"
        return str(x)
    except Exception:
        s = str(x).strip()
        if s in ["0", "negative", "neg", "负向"]:
            return "负向"
        if s in ["1", "positive", "pos", "正向"]:
            return "正向"
        return s


def read_prediction_file(pred_file: Path) -> pd.DataFrame:
    for enc in ["utf-8-sig", "utf-8", "gbk"]:
        try:
            return pd.read_csv(pred_file, encoding=enc)
        except Exception:
            continue
    raise ValueError(f"无法读取文件：{pred_file}")


def collect_one_experiment_errors(
    dataset_display_name: str,
    exp_dir: Path,
    max_errors: int = 5,
) -> pd.DataFrame:
    pred_file = exp_dir / "test_predictions.csv"
    if not pred_file.exists():
        raise FileNotFoundError(f"找不到文件：{pred_file}")

    df = read_prediction_file(pred_file)

    text_col = detect_text_column(df)
    true_col = detect_true_label_column(df)
    pred_col = detect_pred_label_column(df)
    conf_col = detect_confidence_column(df)

    error_df = df[df[true_col].astype(str) != df[pred_col].astype(str)].copy()

    if len(error_df) == 0:
        return pd.DataFrame(columns=[
            "数据集", "模型", "文本", "真实标签", "预测标签", "置信度", "错误类型（待填写）"
        ])

    # 关键修复：重置索引，避免只有第一行有值
    error_df = error_df.head(max_errors).reset_index(drop=True).copy()

    model_display_name = MODEL_NAME_MAP.get(exp_dir.name, exp_dir.name)

    result = pd.DataFrame({
        "数据集": [dataset_display_name] * len(error_df),
        "模型": [model_display_name] * len(error_df),
        "文本": error_df[text_col].astype(str).fillna("").tolist(),
        "真实标签": error_df[true_col].apply(normalize_label).tolist(),
        "预测标签": error_df[pred_col].apply(normalize_label).tolist(),
    })

    if conf_col is not None:
        result["置信度"] = pd.to_numeric(error_df[conf_col], errors="coerce").round(4).tolist()
    else:
        result["置信度"] = [""] * len(error_df)

    result["错误类型（待填写）"] = [""] * len(error_df)
    return result


def beautify_excel(xlsx_path: Path):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # 表头
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 内容
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column == 3:  # 文本列
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 列宽
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 65
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 22

    wb.save(xlsx_path)


def main():
    all_errors = []

    for dataset_display_name, folder_name in BEST_MODEL_BY_DATASET.items():
        exp_dir = MODELS_DIR / folder_name
        if not exp_dir.exists():
            print(f"警告：目录不存在，跳过 -> {exp_dir}")
            continue

        try:
            error_df = collect_one_experiment_errors(
                dataset_display_name=dataset_display_name,
                exp_dir=exp_dir,
                max_errors=5,
            )

            out_csv = OUTPUT_DIR / f"{dataset_display_name}_error_cases.csv"
            out_xlsx = OUTPUT_DIR / f"{dataset_display_name}_error_cases.xlsx"

            error_df.to_csv(out_csv, index=False, encoding="utf-8-sig")
            error_df.to_excel(out_xlsx, index=False)
            beautify_excel(out_xlsx)

            print(f"已生成：{out_xlsx}")
            all_errors.append(error_df)

        except Exception as e:
            print(f"处理失败：{exp_dir}")
            print("原因：", e)

    if all_errors:
        merged_df = pd.concat(all_errors, ignore_index=True)

        merged_csv = OUTPUT_DIR / "all_error_cases_new.csv"
        merged_xlsx = OUTPUT_DIR / "all_error_cases_new.xlsx"

        merged_df.to_csv(merged_csv, index=False, encoding="utf-8-sig")
        merged_df.to_excel(merged_xlsx, index=False)
        beautify_excel(merged_xlsx)

        print("\n===== 错误案例汇总完成 =====")
        print("总CSV：", merged_csv)
        print("总Excel：", merged_xlsx)
    else:
        print("没有收集到任何错误案例。")


if __name__ == "__main__":
    main()