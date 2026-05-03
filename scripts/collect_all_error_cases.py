from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

PROJECT_DIR = Path(r"D:\Users\30126\Desktop\project")
MODELS_DIR = PROJECT_DIR / "models"
OUTPUT_DIR = PROJECT_DIR / "error_case_analysis"
OUTPUT_DIR.mkdir(exist_ok=True)

BEST_MODEL_BY_DATASET = {
    "ChnSentiCorp": "bert_chnsenticorp",
    "OnlineShopping10Cats": "bert_bigru_online",
    "WeiboSenti100k": "roberta_weibo",
}

MODEL_NAME_MAP = {
    "bert_chnsenticorp": "BERT",
    "bert_bigru_online": "BERT+BiGRU",
    "roberta_weibo": "RoBERTa-wwm-ext",
}

def detect_text_column(df: pd.DataFrame) -> str:
    for col in ["text", "sentence", "content", "review", "comment"]:
        if col in df.columns:
            return col
    for col in df.columns:
        if df[col].dtype == "object":
            return col
    raise ValueError("找不到文本列")

def detect_true_label_column(df: pd.DataFrame) -> str:
    for col in ["label", "true_label", "gold_label", "y_true"]:
        if col in df.columns:
            return col
    raise ValueError("找不到真实标签列")

def detect_pred_label_column(df: pd.DataFrame) -> str:
    for col in ["pred_label", "prediction", "pred", "y_pred", "pred_label_id"]:
        if col in df.columns:
            return col
    raise ValueError("找不到预测标签列")

def detect_confidence_column(df: pd.DataFrame) -> str | None:
    for col in ["confidence", "score", "prob", "probability"]:
        if col in df.columns:
            return col
    return None

def normalize_label(x):
    try:
        xi = int(float(x))
        return "负向" if xi == 0 else "正向" if xi == 1 else str(x)
    except Exception:
        s = str(x).strip()
        if s in ["0", "negative", "neg", "负向"]:
            return "负向"
        if s in ["1", "positive", "pos", "正向"]:
            return "正向"
        return s

def read_prediction_file(pred_file: Path) -> pd.DataFrame:
    for enc in ["utf-8-sig", "utf-8", "gbk"]:
        try:
            return pd.read_csv(pred_file, encoding=enc)
        except Exception:
            pass
    raise ValueError(f"无法读取文件：{pred_file}")

def collect_one_experiment_errors(dataset_display_name: str, exp_dir: Path) -> pd.DataFrame:
    pred_file = exp_dir / "test_predictions.csv"
    if not pred_file.exists():
        raise FileNotFoundError(f"找不到文件：{pred_file}")

    df = read_prediction_file(pred_file)
    text_col = detect_text_column(df)
    true_col = detect_true_label_column(df)
    pred_col = detect_pred_label_column(df)
    conf_col = detect_confidence_column(df)

    error_df = df[df[true_col].astype(str) != df[pred_col].astype(str)].copy().reset_index(drop=True)

    model_display_name = MODEL_NAME_MAP.get(exp_dir.name, exp_dir.name)

    result = pd.DataFrame({
        "数据集": [dataset_display_name] * len(error_df),
        "模型": [model_display_name] * len(error_df),
        "文本": error_df[text_col].astype(str).fillna("").tolist(),
        "真实标签": error_df[true_col].apply(normalize_label).tolist(),
        "预测标签": error_df[pred_col].apply(normalize_label).tolist(),
    })

    if conf_col is not None:
        result["置信度"] = pd.to_numeric(error_df[conf_col], errors="coerce").round(4).tolist()
    else:
        result["置信度"] = [""] * len(error_df)

    return result

def beautify_excel(xlsx_path: Path):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column == 3:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    wb.save(xlsx_path)

def main():
    all_errors = []
    for dataset_display_name, folder_name in BEST_MODEL_BY_DATASET.items():
        exp_dir = MODELS_DIR / folder_name
        if not exp_dir.exists():
            print(f"警告：目录不存在，跳过 -> {exp_dir}")
            continue

        error_df = collect_one_experiment_errors(dataset_display_name, exp_dir)
        out_xlsx = OUTPUT_DIR / f"{dataset_display_name}_all_error_cases.xlsx"
        out_csv = OUTPUT_DIR / f"{dataset_display_name}_all_error_cases.csv"

        error_df.to_csv(out_csv, index=False, encoding="utf-8-sig")
        error_df.to_excel(out_xlsx, index=False)
        beautify_excel(out_xlsx)
        print(f"{dataset_display_name} 错误样本数: {len(error_df)}")
        print(f"已生成: {out_xlsx}")
        all_errors.append(error_df)

    if all_errors:
        merged_df = pd.concat(all_errors, ignore_index=True)
        merged_xlsx = OUTPUT_DIR / "all_error_cases_full.xlsx"
        merged_csv = OUTPUT_DIR / "all_error_cases_full.csv"
        merged_df.to_csv(merged_csv, index=False, encoding="utf-8-sig")
        merged_df.to_excel(merged_xlsx, index=False)
        beautify_excel(merged_xlsx)
        print(f"\n总错误样本数: {len(merged_df)}")
        print(f"总Excel: {merged_xlsx}")

if __name__ == "__main__":
    main()